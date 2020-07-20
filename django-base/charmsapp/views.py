from django.shortcuts import render
from django.http import HttpResponse, HttpResponseRedirect
from django_tables2 import RequestConfig
import openpyxl

from django.contrib import messages
from django.urls import reverse
from .resources import SampleDataResource
from tablib import Dataset
from .models import SampleData
from .tables import SampleDataTable

from django.core import serializers
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt

# Create your views here.
# Creating an index view...this is the initial view when 
# the user types in localhost:8000/charmsapp/
# This view will be where the excel file is uploaded
def index(request):
    if "GET" == request.method:
        return render(request, 'charmsapp/index.html', {})
        

# This is the upload view
# # the user types in localhost:8000/charmsapp/upload
def upload(request):
    if "GET" == request.method:
        return render(request, 'charmsapp/upload.html', {})
    else:
        sample_resource = SampleDataResource()
        dataset = Dataset()

        #excel_file = ""
        excel_file = request.FILES["excel_file"]
        file_name = request.FILES['excel_file'].name
        print (file_name)

        # error check to make sure the file is an excel file
        if not excel_file.name.endswith('xlsx'):
            messages.info(request, 'File has the wrong format, must be an xlsx file!')
            return render(request, 'charmsapp/upload.html', {})
        
        # Loads an excel file
        imported_data = dataset.load(excel_file.read(), format='xlsx')
        #print (imported_data)

        # do data validation
        # first check to see if we have any data in our database yet
        if SampleData.objects.exists():
            data_headers = [f.get_attname() for f in SampleData._meta.get_fields()]
            """
            Testing a theory here
            object_list = serializers.serialize("python", SampleData.objects.all())
            for object in object_list:
                for field_name, field_value in object['fields'].items():
                    print (field_name, field_value)
            """
            data_count = 0
            for data in imported_data:
                if "Total" not in data[0]:
                    data_count = data_count + 1
                    """
                    Question: any way to make the items that I filter on more dynaic?  Instead of having to type
                    them up for everytime they might change?
                    """
                    filter_results = SampleData.objects.filter(id = data_count,
                            first_name = data[0],
                            last_name = data[1],
                            email = data[2],
                            fav_number = data[3],
                            sample_date = data[4])
                    if not filter_results.exists():
                        print (" Updating the database with new values ")
                        value = SampleData(
                            id = data_count,
                            first_name = data[0],
                            last_name = data[1],
                            email = data[2],
                            fav_number = data[3],
                            sample_date = data[4],
                            )
                        value.save()
        else:
            # This means this is our first time inserting into the database
            print ("First time importing data")
            data_count = 0
            for data in imported_data:
                print("Data data did not contain total so we are saving it ")
                if "Total" not in data[0]:
                    data_count = data_count + 1
                    value = SampleData(
                        id = data_count,
                        first_name = data[0],
                        last_name = data[1],
                        email = data[2],
                        fav_number = data[3],
                        sample_date = data[4],
                    )
                    value.save()

        """
        # you may put validations here to check extension or file size
        wb = openpyxl.load_workbook(excel_file)

        # getting all sheets
        sheets = wb.sheetnames
        #print(sheets)

        # getting a particular sheet
        worksheet = wb["Sheet1"]
        #print(worksheet)

        # getting active sheet
        active_sheet = wb.active
        #print(active_sheet)

        # reading a cell
        #print(worksheet["A2"].value)

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        #for row in worksheet.iter_rows():
        #    row_data = list()
        #    for cell in row:
        #        row_data.append(str(cell.value))
        #        print(cell.value)
        #    excel_data.append(row_data)
        """

        return render(request, 'charmsapp/upload.html', {"excel_data":imported_data})

@csrf_exempt
def review_data(request):
    if "GET" == request.method:
        # return HttpResponse("GET REQUEST: This is the view/review page");
        #sample_data = SampleDataTable(SampleData.objects.all())
        #RequestConfig(request).configure(sample_data)
        sample_data = SampleData.objects.all()
        data_headers = [f.get_attname() for f in SampleData._meta.get_fields()]
        #print (data_headers)
        #for header in data_headers:
        #print (header)
        return render(request, "charmsapp/review.html", {
        'test_data':sample_data,
        'headers': data_headers
        })
    else:
        return HttpResponse("POST REQUEST: This is the view/review page");
        
@csrf_exempt
def update_data(request):
        data_id=request.POST.get('id','')
        print (data_id)
        data_type=request.POST.get('type','')
        print (data_type)
        value=request.POST.get('value','')
        print (value)

        testData=SampleData.objects.get(id=data_id)
        print (testData.first_name)

        if data_type=="first_name":
           testData.first_name=value    
    
        if data_type=="last_name":
           testData.last_name=value

        if data_type == "email":
            testData.email = value

        if data_type == "fav_number":
            testData.fav_number = value

        if data_type == "sample_date":
            testData.sample_date = value

        print (testData.first_name)

        testData.save()
        return JsonResponse({"success":"Updated"})

@csrf_exempt
def delete_data(request):
    data_id=request.POST.get('id','')
    testData=SampleData.objects.get(id=data_id)
    testData.delete()
    #messages.error(request, "Deleted Successfully")
    return JsonResponse({"success":f"Succesfully deleted row {data_id}"})