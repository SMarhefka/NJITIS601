from django.http import HttpResponse
from django.shortcuts import render
import openpyxl

from django.contrib import messages
from .resources import SampleDataResource
from tablib import Dataset
from .models import SampleExcelData
from .forms import SampleFormData

def index(request):
    return HttpResponse("This is currently the index or the landing page!");

def upload_1(request):
    if "GET" == request.method:
        return render(request, reverse('admin:app_list', kwargs={'charmsapp': 'upload'}), {})
    else:
        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # getting all sheets
        sheets = wb.sheetnames
        print(sheets)

        # getting a particular sheet
        worksheet = wb["Sheet1"]
        print(worksheet)

        # getting active sheet
        active_sheet = wb.active
        print(active_sheet)

        # reading a cell
        print(worksheet["A1"].value)

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
                print(cell.value)
            excel_data.append(row_data)

        return render(request, 'sampleExcel/upload.html', {"excel_data":excel_data})

def upload(request):
    if request.method == 'POST':
        form = SampleFormData(request.POST,request.FILES)
        if form.is_valid():
            sample_resource = SampleDataResource()
            dataset = Dataset()
            excel_file = ""
            excel_file = request.FILES['file']
            # Loads an excel file
            imported_data = dataset.load(excel_file.read(), format='xlsx')
            print (imported_data)
            count=0
            for data in imported_data:
        	    # print(data[1])
        	    value = SampleExcelData(
                    first_name = data[0],
                    last_name = data[1],
                    email = data[2],
                    fav_number = data[3],
                    date = data[4],
        		    )
        	    value.save()
            return render(request, 'sampleExcel/upload.html', {"excel_data":imported_data})
        else:
            return render(request, 'sampleExcel/upload.html', {"form":form})
    else:
        form = SampleFormData()
        return render(request, 'sampleExcel/upload.html', {"form":form})
    